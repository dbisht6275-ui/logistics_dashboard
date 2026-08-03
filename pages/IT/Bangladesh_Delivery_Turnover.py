import io
from datetime import date, datetime, timedelta

import pandas as pd
import streamlit as st
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from services.database import get_engine


# =========================================================
# CONSTANTS
# =========================================================

TEXT_COLUMNS = [
    "ZONENAME",
    "HUBNAME",
    "BRANCH",
]


# =========================================================
# DATE HELPERS
# =========================================================

def validate_date_range(
    period_name: str,
    from_date: date,
    to_date: date,
) -> None:
    """
    Validate one selected date range.
    """

    if from_date > to_date:
        raise ValueError(
            f"{period_name}: From Date cannot be after To Date."
        )


def format_period_column(
    period_number: int,
    from_date: datetime,
    to_date: datetime,
    is_ftl: bool = False,
) -> str:
    """
    Create dynamic column names.

    Examples:
        1. MAY-26 TO JUL-26
        2. JUL-26
        3. AUG-26 (FTL)
    """

    from_month = from_date.strftime("%b-%y").upper()
    to_month = to_date.strftime("%b-%y").upper()

    same_month = (
        from_date.year == to_date.year
        and from_date.month == to_date.month
    )

    if same_month:
        heading = to_month
    else:
        heading = f"{from_month} TO {to_month}"

    if is_ftl:
        heading += " (FTL)"

    return f"{period_number}. {heading}"


# =========================================================
# DATABASE REPORT FUNCTION
# =========================================================

def get_bangladesh_delivery_turnover(
    date1_from: str,
    date1_to: str,
    date2_from: str,
    date2_to: str,
    date3_from: str,
    date3_to: str,
):
    """
    Get branch-wise Bangladesh delivery turnover.

    Input date format:
        DD-MM-YYYY

    Return:
        columns, rows
    """

    connection = None

    try:
        # -------------------------------------------------
        # DATE CONVERSION
        # -------------------------------------------------

        from1_dt = datetime.strptime(
            date1_from,
            "%d-%m-%Y",
        )

        to1_dt = datetime.strptime(
            date1_to,
            "%d-%m-%Y",
        )

        from2_dt = datetime.strptime(
            date2_from,
            "%d-%m-%Y",
        )

        to2_dt = datetime.strptime(
            date2_to,
            "%d-%m-%Y",
        )

        from3_dt = datetime.strptime(
            date3_from,
            "%d-%m-%Y",
        )

        to3_dt = datetime.strptime(
            date3_to,
            "%d-%m-%Y",
        )

        validate_date_range(
            "Period 1",
            from1_dt.date(),
            to1_dt.date(),
        )

        validate_date_range(
            "Period 2",
            from2_dt.date(),
            to2_dt.date(),
        )

        validate_date_range(
            "Period 3",
            from3_dt.date(),
            to3_dt.date(),
        )

        # Exclusive end date:
        # CN.GRDT >= start AND CN.GRDT < next day
        # This includes the complete selected To Date.
        to1_exclusive = to1_dt + timedelta(days=1)
        to2_exclusive = to2_dt + timedelta(days=1)
        to3_exclusive = to3_dt + timedelta(days=1)

        # -------------------------------------------------
        # DYNAMIC COLUMN NAMES
        # -------------------------------------------------

        col1_non_ftl = format_period_column(
            1,
            from1_dt,
            to1_dt,
            is_ftl=False,
        )

        col2_non_ftl = format_period_column(
            2,
            from2_dt,
            to2_dt,
            is_ftl=False,
        )

        col3_non_ftl = format_period_column(
            3,
            from3_dt,
            to3_dt,
            is_ftl=False,
        )

        col1_ftl = format_period_column(
            1,
            from1_dt,
            to1_dt,
            is_ftl=True,
        )

        col2_ftl = format_period_column(
            2,
            from2_dt,
            to2_dt,
            is_ftl=True,
        )

        col3_ftl = format_period_column(
            3,
            from3_dt,
            to3_dt,
            is_ftl=True,
        )

        # -------------------------------------------------
        # DATABASE CONNECTION
        # -------------------------------------------------

        engine = get_engine()
        connection = engine.raw_connection()

        # -------------------------------------------------
        # SQL QUERY
        # -------------------------------------------------

        query = f"""
            SELECT
                ZONE.ZONENAME,
                ZONE.HUBNAME,
                ORG.STNNAME AS BRANCH,

                SUM(
                    CASE
                        WHEN ISNULL(CN.FTL, 'N') <> 'Y'
                             AND CN.GRDT >= ?
                             AND CN.GRDT < ?
                        THEN (
                            ISNULL(CN.TAMOUNT, 0)
                            - ISNULL(CN.SERVICETAX, 0)
                        ) / 300000.0
                        ELSE 0
                    END
                ) AS [{col1_non_ftl}],

                SUM(
                    CASE
                        WHEN ISNULL(CN.FTL, 'N') <> 'Y'
                             AND CN.GRDT >= ?
                             AND CN.GRDT < ?
                        THEN (
                            ISNULL(CN.TAMOUNT, 0)
                            - ISNULL(CN.SERVICETAX, 0)
                        ) / 100000.0
                        ELSE 0
                    END
                ) AS [{col2_non_ftl}],

                SUM(
                    CASE
                        WHEN ISNULL(CN.FTL, 'N') <> 'Y'
                             AND CN.GRDT >= ?
                             AND CN.GRDT < ?
                        THEN (
                            ISNULL(CN.TAMOUNT, 0)
                            - ISNULL(CN.SERVICETAX, 0)
                        ) / 100000.0
                        ELSE 0
                    END
                ) AS [{col3_non_ftl}],

                SUM(
                    CASE
                        WHEN ISNULL(CN.FTL, 'N') = 'Y'
                             AND CN.GRDT >= ?
                             AND CN.GRDT < ?
                        THEN (
                            ISNULL(CN.TAMOUNT, 0)
                            - ISNULL(CN.SERVICETAX, 0)
                        ) / 300000.0
                        ELSE 0
                    END
                ) AS [{col1_ftl}],

                SUM(
                    CASE
                        WHEN ISNULL(CN.FTL, 'N') = 'Y'
                             AND CN.GRDT >= ?
                             AND CN.GRDT < ?
                        THEN (
                            ISNULL(CN.TAMOUNT, 0)
                            - ISNULL(CN.SERVICETAX, 0)
                        ) / 100000.0
                        ELSE 0
                    END
                ) AS [{col2_ftl}],

                SUM(
                    CASE
                        WHEN ISNULL(CN.FTL, 'N') = 'Y'
                             AND CN.GRDT >= ?
                             AND CN.GRDT < ?
                        THEN (
                            ISNULL(CN.TAMOUNT, 0)
                            - ISNULL(CN.SERVICETAX, 0)
                        ) / 100000.0
                        ELSE 0
                    END
                ) AS [{col3_ftl}]

            FROM CNMT CN WITH (NOLOCK)

            INNER JOIN STATIONMAST ORG
                ON ORG.STNCODE = CN.ORGCODE

            INNER JOIN VIEWSTATIONMAST ZONE
                ON ZONE.STNCODE = CN.ORGCODE

            INNER JOIN STATIONMAST DST
                ON DST.STNCODE = CN.DESTCODE

            WHERE
                CN.GRTYPE <> 'N'

                AND (
                    UPPER(
                        LTRIM(
                            RTRIM(
                                ISNULL(DST.COUNTRY, '')
                            )
                        )
                    ) = 'BANGLADESH'

                    OR UPPER(
                        LTRIM(
                            RTRIM(
                                ISNULL(DST.STNNAME, '')
                            )
                        )
                    ) IN (
                        'PETRAPOLE',
                        'MYMENSINGH'
                    )
                )

                AND (
                    (
                        CN.GRDT >= ?
                        AND CN.GRDT < ?
                    )
                    OR
                    (
                        CN.GRDT >= ?
                        AND CN.GRDT < ?
                    )
                    OR
                    (
                        CN.GRDT >= ?
                        AND CN.GRDT < ?
                    )
                )

            GROUP BY
                ZONE.ZONENAME,
                ZONE.HUBNAME,
                ORG.STNNAME

            ORDER BY
                ZONE.ZONENAME,
                ZONE.HUBNAME,
                ORG.STNNAME
        """

        parameters = [
            # Non-FTL Period 1
            from1_dt,
            to1_exclusive,

            # Non-FTL Period 2
            from2_dt,
            to2_exclusive,

            # Non-FTL Period 3
            from3_dt,
            to3_exclusive,

            # FTL Period 1
            from1_dt,
            to1_exclusive,

            # FTL Period 2
            from2_dt,
            to2_exclusive,

            # FTL Period 3
            from3_dt,
            to3_exclusive,

            # Main WHERE Period 1
            from1_dt,
            to1_exclusive,

            # Main WHERE Period 2
            from2_dt,
            to2_exclusive,

            # Main WHERE Period 3
            from3_dt,
            to3_exclusive,
        ]

        dataframe = pd.read_sql_query(
            query,
            connection,
            params=parameters,
        )

        if dataframe.empty:
            return [], []

        # -------------------------------------------------
        # CLEAN DATABASE RESULT
        # -------------------------------------------------

        for column in TEXT_COLUMNS:
            if column in dataframe.columns:
                dataframe[column] = (
                    dataframe[column]
                    .fillna("Unknown")
                    .astype(str)
                    .str.strip()
                    .replace("", "Unknown")
                )

        numeric_columns = [
            column
            for column in dataframe.columns
            if column not in TEXT_COLUMNS
        ]

        for column in numeric_columns:
            dataframe[column] = pd.to_numeric(
                dataframe[column],
                errors="coerce",
            ).fillna(0.0)

        dataframe[numeric_columns] = (
            dataframe[numeric_columns].round(2)
        )

        return (
            list(dataframe.columns),
            dataframe.values.tolist(),
        )

    except Exception as error:
        print(
            "Error in Bangladesh Delivery Turnover:",
            error,
        )
        return [], []

    finally:
        if connection is not None:
            try:
                connection.close()
            except Exception:
                pass


# =========================================================
# GRAND TOTAL ROW
# =========================================================

def add_grand_total_row(
    dataframe: pd.DataFrame,
) -> pd.DataFrame:
    """
    Add a grand-total row at the bottom.
    """

    report_dataframe = dataframe.copy()

    numeric_columns = [
        column
        for column in report_dataframe.columns
        if column not in TEXT_COLUMNS
    ]

    for column in numeric_columns:
        report_dataframe[column] = pd.to_numeric(
            report_dataframe[column],
            errors="coerce",
        ).fillna(0.0)

    total_row = {
        column: ""
        for column in report_dataframe.columns
    }

    if "ZONENAME" in total_row:
        total_row["ZONENAME"] = "GRAND TOTAL"

    for column in numeric_columns:
        total_row[column] = round(
            float(report_dataframe[column].sum()),
            2,
        )

    total_dataframe = pd.DataFrame(
        [total_row]
    )

    return pd.concat(
        [
            report_dataframe,
            total_dataframe,
        ],
        ignore_index=True,
    )


# =========================================================
# EXCEL EXPORT
# =========================================================

def create_bangladesh_turnover_excel(
    report_dataframe: pd.DataFrame,
) -> bytes:
    """
    Create an Excel report in memory.
    """

    output = io.BytesIO()

    with pd.ExcelWriter(
        output,
        engine="openpyxl",
    ) as writer:

        report_dataframe.to_excel(
            writer,
            index=False,
            sheet_name="Bangladesh Turnover",
        )

        worksheet = writer.sheets[
            "Bangladesh Turnover"
        ]

        worksheet.freeze_panes = "D2"
        worksheet.auto_filter.ref = worksheet.dimensions
        worksheet.sheet_view.showGridLines = False

        header_fill = PatternFill(
            fill_type="solid",
            fgColor="1F4E78",
        )

        header_font = Font(
            bold=True,
            color="FFFFFF",
        )

        total_fill = PatternFill(
            fill_type="solid",
            fgColor="D9EAF7",
        )

        total_font = Font(
            bold=True,
            color="17365D",
        )

        # Header formatting
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

        worksheet.row_dimensions[1].height = 38

        # Data formatting
        for row_number in range(
            2,
            worksheet.max_row + 1,
        ):
            for column_number in range(
                1,
                worksheet.max_column + 1,
            ):
                cell = worksheet.cell(
                    row=row_number,
                    column=column_number,
                )

                if column_number >= 4:
                    cell.number_format = "0.00"

                cell.alignment = Alignment(
                    vertical="center",
                    wrap_text=False,
                )

        # Grand-total formatting
        total_row_number = worksheet.max_row

        for cell in worksheet[total_row_number]:
            cell.fill = total_fill
            cell.font = total_font

        # Column widths
        for column_number, column_name in enumerate(
            report_dataframe.columns,
            start=1,
        ):
            column_letter = get_column_letter(
                column_number
            )

            maximum_length = len(str(column_name))

            for row_number in range(
                2,
                worksheet.max_row + 1,
            ):
                value = worksheet.cell(
                    row=row_number,
                    column=column_number,
                ).value

                if value is not None:
                    maximum_length = max(
                        maximum_length,
                        len(str(value)),
                    )

            if column_name == "ZONENAME":
                width = 20

            elif column_name == "HUBNAME":
                width = 22

            elif column_name == "BRANCH":
                width = 26

            else:
                width = min(
                    max(maximum_length + 3, 16),
                    26,
                )

            worksheet.column_dimensions[
                column_letter
            ].width = width

    output.seek(0)

    return output.getvalue()


# =========================================================
# STREAMLIT UI
# =========================================================

def show_bangladesh_delivery_turnover():
    """
    Display Bangladesh Delivery Turnover report.
    """

    st.markdown(
        """
        <div style="
            padding:14px 18px;
            margin-bottom:12px;
            border:1px solid #d8e2f0;
            border-radius:12px;
            background:linear-gradient(
                180deg,
                #ffffff 0%,
                #f7faff 100%
            );
        ">
            <div style="
                color:#17365d;
                font-size:21px;
                font-weight:750;
            ">
                Bangladesh Delivery Turnover
            </div>

            <div style="
                margin-top:3px;
                color:#64748b;
                font-size:12px;
            ">
                Branch-wise Non-FTL and FTL delivery
                turnover comparison
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    # -----------------------------------------------------
    # DEFAULT DATES
    # -----------------------------------------------------

    today = date.today()

    current_month_start = today.replace(
        day=1
    )

    previous_month_end = (
        current_month_start
        - timedelta(days=1)
    )

    previous_month_start = (
        previous_month_end.replace(
            day=1
        )
    )

    period1_end = (
        previous_month_start
        - timedelta(days=1)
    )

    # Period 1 default:
    # previous three complete months
    period1_start_month = (
        period1_end.month - 2
    )

    period1_start_year = (
        period1_end.year
    )

    while period1_start_month <= 0:
        period1_start_month += 12
        period1_start_year -= 1

    period1_start = date(
        period1_start_year,
        period1_start_month,
        1,
    )

    # -----------------------------------------------------
    # FILTER UI
    # -----------------------------------------------------

    with st.container(border=True):

        st.markdown(
            """
            <div style="
                margin-bottom:8px;
                color:#17365d;
                font-size:15px;
                font-weight:700;
            ">
                Select Report Periods
            </div>
            """,
            unsafe_allow_html=True,
        )

        period1_column, period2_column, period3_column = (
            st.columns(
                3,
                gap="medium",
            )
        )

        with period1_column:

            st.markdown("**Period 1**")

            date1_from = st.date_input(
                "From Date",
                value=period1_start,
                format="DD-MM-YYYY",
                key="bangladesh_period1_from",
            )

            date1_to = st.date_input(
                "To Date",
                value=period1_end,
                format="DD-MM-YYYY",
                key="bangladesh_period1_to",
            )

        with period2_column:

            st.markdown("**Period 2**")

            date2_from = st.date_input(
                "From Date",
                value=previous_month_start,
                format="DD-MM-YYYY",
                key="bangladesh_period2_from",
            )

            date2_to = st.date_input(
                "To Date",
                value=previous_month_end,
                format="DD-MM-YYYY",
                key="bangladesh_period2_to",
            )

        with period3_column:

            st.markdown("**Period 3**")

            date3_from = st.date_input(
                "From Date",
                value=current_month_start,
                format="DD-MM-YYYY",
                key="bangladesh_period3_from",
            )

            date3_to = st.date_input(
                "To Date",
                value=today,
                format="DD-MM-YYYY",
                key="bangladesh_period3_to",
            )

        generate_report = st.button(
            "Generate Report",
            type="primary",
            use_container_width=True,
            key="generate_bangladesh_delivery_report",
        )

    # Save dates after Generate Report is clicked.
    # This keeps the report visible after download interactions.
    if generate_report:

        st.session_state[
            "bangladesh_delivery_report_dates"
        ] = {
            "date1_from": date1_from,
            "date1_to": date1_to,
            "date2_from": date2_from,
            "date2_to": date2_to,
            "date3_from": date3_from,
            "date3_to": date3_to,
        }

    saved_dates = st.session_state.get(
        "bangladesh_delivery_report_dates"
    )

    if not saved_dates:
        return

    date1_from = saved_dates["date1_from"]
    date1_to = saved_dates["date1_to"]

    date2_from = saved_dates["date2_from"]
    date2_to = saved_dates["date2_to"]

    date3_from = saved_dates["date3_from"]
    date3_to = saved_dates["date3_to"]

    # -----------------------------------------------------
    # DATE VALIDATION
    # -----------------------------------------------------

    try:
        validate_date_range(
            "Period 1",
            date1_from,
            date1_to,
        )

        validate_date_range(
            "Period 2",
            date2_from,
            date2_to,
        )

        validate_date_range(
            "Period 3",
            date3_from,
            date3_to,
        )

    except ValueError as error:
        st.error(str(error))
        return

    # -----------------------------------------------------
    # LOAD REPORT DATA
    # -----------------------------------------------------

    try:
        with st.spinner(
            "Loading Bangladesh delivery turnover report..."
        ):
            columns, rows = (
                get_bangladesh_delivery_turnover(
                    date1_from.strftime(
                        "%d-%m-%Y"
                    ),
                    date1_to.strftime(
                        "%d-%m-%Y"
                    ),
                    date2_from.strftime(
                        "%d-%m-%Y"
                    ),
                    date2_to.strftime(
                        "%d-%m-%Y"
                    ),
                    date3_from.strftime(
                        "%d-%m-%Y"
                    ),
                    date3_to.strftime(
                        "%d-%m-%Y"
                    ),
                )
            )

    except Exception as error:
        st.error(
            "Unable to generate Bangladesh "
            "Delivery Turnover report."
        )
        st.exception(error)
        return

    if not columns or not rows:
        st.warning(
            "No Bangladesh delivery turnover data "
            "was found for the selected periods."
        )
        return

    dataframe = pd.DataFrame(
        rows,
        columns=columns,
    )

    # -----------------------------------------------------
    # CLEAN UI DATA
    # -----------------------------------------------------

    numeric_columns = [
        column
        for column in dataframe.columns
        if column not in TEXT_COLUMNS
    ]

    for column in TEXT_COLUMNS:
        if column in dataframe.columns:
            dataframe[column] = (
                dataframe[column]
                .fillna("Unknown")
                .astype(str)
                .str.strip()
                .replace("", "Unknown")
            )

    for column in numeric_columns:
        dataframe[column] = pd.to_numeric(
            dataframe[column],
            errors="coerce",
        ).fillna(0.0).round(2)

    report_dataframe = add_grand_total_row(
        dataframe
    )

    # -----------------------------------------------------
    # TABLE COLUMN CONFIGURATION
    # -----------------------------------------------------

    column_configuration = {}

    if "ZONENAME" in report_dataframe.columns:

        column_configuration[
            "ZONENAME"
        ] = st.column_config.TextColumn(
            "Zone",
            width="medium",
        )

    if "HUBNAME" in report_dataframe.columns:

        column_configuration[
            "HUBNAME"
        ] = st.column_config.TextColumn(
            "Hub",
            width="medium",
        )

    if "BRANCH" in report_dataframe.columns:

        column_configuration[
            "BRANCH"
        ] = st.column_config.TextColumn(
            "Branch",
            width="large",
        )

    for column in numeric_columns:

        column_configuration[
            column
        ] = st.column_config.NumberColumn(
            column,
            format="%.2f",
            width="small",
        )

    # -----------------------------------------------------
    # REPORT TABLE
    # -----------------------------------------------------

    st.markdown(
        """
        <div style="
            margin:15px 0 8px;
            color:#17365d;
            font-size:17px;
            font-weight:700;
        ">
            Bangladesh Delivery Turnover Report
        </div>
        """,
        unsafe_allow_html=True,
    )

    st.dataframe(
        report_dataframe,
        use_container_width=True,
        height=560,
        hide_index=True,
        column_config=column_configuration,
    )

    st.caption(
        "Values are displayed in ₹ lakh. "
        "Period 1 uses the existing division by 300,000. "
        "Periods 2 and 3 use division by 100,000."
    )

    # -----------------------------------------------------
    # DOWNLOADS
    # -----------------------------------------------------

    excel_data = create_bangladesh_turnover_excel(
        report_dataframe
    )

    csv_data = report_dataframe.to_csv(
        index=False,
    ).encode("utf-8-sig")

    excel_column, csv_column = st.columns(
        2
    )

    with excel_column:

        st.download_button(
            label="Download Excel",
            data=excel_data,
            file_name=(
                "bangladesh_delivery_turnover_"
                f"{date1_from:%Y%m%d}_"
                f"{date3_to:%Y%m%d}.xlsx"
            ),
            mime=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
            use_container_width=True,
            key="download_bangladesh_excel",
        )

    with csv_column:

        st.download_button(
            label="Download CSV",
            data=csv_data,
            file_name=(
                "bangladesh_delivery_turnover_"
                f"{date1_from:%Y%m%d}_"
                f"{date3_to:%Y%m%d}.csv"
            ),
            mime="text/csv",
            use_container_width=True,
            key="download_bangladesh_csv",
        )


# =========================================================
# OPTIONAL ALIAS
# =========================================================

def show_bangladesh_turnover():
    show_bangladesh_delivery_turnover()
